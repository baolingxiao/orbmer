#!/usr/bin/env python3
"""Import all Jinqi (今奇) toy products + images from the supplier xlsx."""

from __future__ import annotations

import json
import re
import shutil
from io import BytesIO
from pathlib import Path

import openpyxl
from PIL import Image

XLSX = Path(
    "/Users/dai/Library/Containers/com.tencent.xinWeChat/Data/Library/Caches/"
    "com.tencent.xinWeChat/2.0b4.0.9/5cfac9ee3bb10fed59840cccea23ffb4/SaveTemp/"
    "22fd500cc15363b5581c215e02ac69ab/今奇.xlsx"
)
ROOT = Path(__file__).resolve().parents[1]
OUT_IMG = ROOT / "web" / "assets" / "toys" / "jinqi"
OUT_JSON = ROOT / "data" / "jinqi-catalog.json"
OUT_JS = ROOT / "web" / "shared" / "js" / "catalog-toys.js"

# Factory CNY → retail USD: ~7.2 FX, ~4× margin
FX = 7.2
MARKUP = 4.0

WORD_MAP = [
    ("传统中国龙", "Traditional Chinese Dragon"),
    ("宝石飞行龙（长尾）", "Gem Flying Dragon (Long Tail)"),
    ("翅膀飞行宝石龙（短尾）", "Winged Gem Dragon (Short Tail)"),
    ("BABY龙家族飞龙组合", "Baby Dragon Family Set"),
    ("史莱姆龙蛋（解压龙蛋、扭扭蛋）", "Slime Dragon Egg Fidget"),
    ("解压球（人工组装8个零部件）", "Fidget Sphere (8-part)"),
    ("齿轮球旋转球按压球", "Gear Spin Press Ball"),
    ("笑脸、爱心、轮胎解压盘", "Smile Heart Tire Fidget Disc"),
    ("杰克逊变色龙（角+眼睛是夜光）", "Jackson Chameleon (Glow Accents)"),
    ("夜光骷髅手脚骨架", "Glow Skeleton Hands & Feet"),
    ("按键包子解压钥匙扣", "Bao Button Fidget Keychain"),
    ("按键包子解压玩具", "Bao Button Fidget Toy"),
    ("磁吸娃娃鱼冰箱贴", "Magnetic Axolotl Fridge Magnet"),
    ("磁吸青蛙冰箱贴", "Magnetic Frog Fridge Magnet"),
    ("半眼多色小动物", "Half-Eye Multicolor Critters"),
    ("口袋小动物系列", "Pocket Critter Series"),
    ("奇幻龙蛋水果多色套装", "Fantasy Fruit Dragon Egg Set"),
    ("多色夜光龙蛋套装", "Multicolor Glow Dragon Egg Set"),
    ("冰晶龙蛋套装", "Ice Crystal Dragon Egg Set"),
    ("恐龙龙蛋盒装", "Dino Dragon Egg Box Set"),
    ("水晶角龙套装", "Crystal Horn Dragon Set"),
    ("凤凰龙蛋", "Phoenix Dragon Egg"),
    ("蝴蝶龙蛋", "Butterfly Dragon Egg"),
    ("破壳龙蛋", "Hatching Dragon Egg"),
    ("开合龙蛋", "Hinged Dragon Egg"),
    ("泡泡龙蛋", "Bubble Dragon Egg"),
    ("兔子蛋", "Bunny Egg"),
    ("多色兔子蛋套装", "Multicolor Bunny Egg Set"),
    ("恐龙蛋多色套装", "Multicolor Dino Egg Set"),
    ("多色恐龙骨架", "Multicolor Dino Skeleton"),
    ("迷你霸王龙", "Mini T-Rex"),
    ("宝石龙", "Gem Dragon"),
    ("机甲龙", "Mech Dragon"),
    ("可立龙", "Standing Dragon"),
    ("祥云龙", "Cloud Dragon"),
    ("无牙龙", "Toothless Dragon"),
    ("水晶角龙", "Crystal Horn Dragon"),
    ("龙之家族", "Dragon Family"),
    ("大角龙", "Big Horn Dragon"),
    ("霸王龙", "T-Rex"),
    ("剑龙", "Stegosaurus"),
    ("三角龙", "Triceratops"),
    ("棘背龙", "Spinosaurus"),
    ("迅猛龙", "Velociraptor"),
    ("猛犸象", "Mammoth"),
    ("甲龙骨", "Ankylosaurus Skeleton"),
    ("鳄鱼骨", "Crocodile Skeleton"),
    ("鲨鱼鱼骨", "Shark Skeleton"),
    ("龙蛋盒装", "Dragon Egg Box Set"),
    ("仙女龙", "Fairy Dragon"),
    ("旋风迅猛龙", "Whirl Velociraptor"),
    ("旋风霸王龙", "Whirl T-Rex"),
    ("弹力蜘蛛", "Spring Spider"),
    ("大眼蝎子", "Big-Eye Scorpion"),
    ("赛博朋克螃蟹", "Cyberpunk Crab"),
    ("水晶宝石螃蟹", "Crystal Gem Crab"),
    ("寄居蟹", "Hermit Crab"),
    ("三足虫", "Trilobite"),
    ("八爪鱼", "Octopus"),
    ("多色鱿鱼", "Multicolor Squid"),
    ("鳗鱼骨头", "Eel Skeleton"),
    ("魔鬼鱼", "Manta Ray"),
    ("仿真海龟", "Realistic Sea Turtle"),
    ("蝰蛇盒装", "Viper Box Set"),
    ("蝰蛇", "Viper"),
    ("眼镜蛇", "Cobra"),
    ("疯狂眼镜蛇", "Crazy Cobra"),
    ("蟒蛇", "Python"),
    ("呆萌猫咪", "Cute Cat"),
    ("无毛暹罗猫", "Hairless Siamese Cat"),
    ("呆萌法斗", "Cute French Bulldog"),
    ("三头犬", "Cerberus"),
    ("天使、恶魔猫", "Angel & Demon Cats"),
    ("魔法猫咪", "Magic Cat"),
    ("大耳朵鼠", "Big-Ear Mouse"),
    ("恶魔吉娃娃", "Demon Chihuahua"),
    ("弹簧小狗", "Spring Puppy"),
    ("幽灵猫狗", "Ghost Cat & Dog"),
    ("冰霜怪兽", "Frost Monster"),
    ("小飞象", "Flying Elephant"),
    ("多眼蜘蛛", "Multi-Eye Spider"),
    ("机械姬", "Mech Girl"),
    ("树懒", "Sloth"),
    ("大眼海龟", "Big-Eye Turtle"),
    ("圣诞小浣熊", "Christmas Raccoon"),
    ("鬃狮蜥", "Bearded Dragon"),
    ("松果刺猬", "Pinecone Hedgehog"),
    ("煎饼章鱼", "Pancake Octopus"),
    ("尼斯湖水怪", "Loch Ness Monster"),
    ("骷髅骨架", "Skeleton"),
    ("变色龙", "Chameleon"),
    ("凤头蜥", "Crested Lizard"),
    ("可爱小马", "Cute Pony"),
    ("天使独角兽", "Angel Unicorn"),
    ("不听不看不说", "See Hear Speak No Evil"),
    ("海盗章鱼", "Pirate Octopus"),
    ("美人鱼", "Mermaid"),
    ("圣诞小人", "Christmas Mini Figure"),
    ("骷髅骰子塔", "Skull Dice Tower"),
    ("机械猫头鹰", "Mech Owl"),
    ("动物解压旋旋乐", "Animal Spin Fidget"),
    ("像素解压水果系列", "Pixel Fruit Fidget"),
    ("多边形旋旋乐1代", "Polygon Spin Fidget V1"),
    ("解压器（2代旋旋乐）", "Spin Fidget V2"),
    ("圣诞解压旋旋乐", "Christmas Spin Fidget"),
    ("铠甲旋旋乐", "Armor Spin Fidget"),
    ("三合一解压器", "3-in-1 Fidget"),
    ("蛋壳旋旋乐", "Eggshell Spin Fidget"),
    ("MINI手握按摩棒", "Mini Hand Massager"),
    ("手握按摩棒", "Hand Massager"),
    ("星际解压盘", "Galaxy Fidget Disc"),
    ("伸缩萝卜", "Extendable Radish"),
    ("迷宫蛋", "Maze Egg"),
    ("锁链剑", "Chain Sword"),
    ("指尖解压", "Fingertip Fidget"),
    ("指尖陀螺", "Fidget Spinner"),
    ("萝卜塔", "Radish Tower"),
    ("解压盘", "Fidget Disc"),
    ("按压旋转球", "Press Spin Ball"),
    ("旋旋乐", "Spin Fidget"),
    ("香蕉", "Banana"),
    ("水果按键", "Fruit Button Fidget"),
    ("解压棒棒糖", "Fidget Lollipop"),
    ("edc水果叮叮币", "EDC Fruit Coin"),
    ("骷髅按键系列", "Skull Button Series"),
    ("按键小动物", "Critter Button Fidget"),
    ("植物按键", "Plant Button Fidget"),
    ("按键乌龟", "Turtle Button Fidget"),
    ("球类按键解压", "Ball Button Fidget"),
    ("太空探索", "Space Explorer"),
    ("猫头鹰钥匙扣", "Owl Keychain"),
    ("变色龙钥匙扣", "Chameleon Keychain"),
    ("乌龟钥匙扣", "Turtle Keychain"),
    ("大象钥匙扣", "Elephant Keychain"),
    ("三角龙钥匙扣", "Triceratops Keychain"),
    ("海豚钥匙扣", "Dolphin Keychain"),
    ("驯鹿钥匙扣", "Reindeer Keychain"),
    ("猫狗系列钥匙扣", "Cat & Dog Keychain"),
    ("编织小熊钥匙扣", "Knit Bear Keychain"),
    ("章鱼钥匙扣", "Octopus Keychain"),
    ("青蛙钥匙扣", "Frog Keychain"),
    ("冰霜怪兽钥匙扣", "Frost Monster Keychain"),
    ("龙虾钥匙扣", "Lobster Keychain"),
    ("多色小动物", "Multicolor Critters"),
    ("海洋系列", "Ocean Series"),
    ("株萌系列", "Sprout Series"),
    ("小动物家族", "Critter Family"),
    ("萌粒鸭子", "Mini Duck"),
    ("蜘蛛家族", "Spider Family"),
    ("恐龙萌粒", "Mini Dino"),
    ("神话系列", "Myth Series"),
    ("面包伙伴坐姿小人", "Bread Buddy Sitters"),
    ("圣诞伙伴坐姿小人", "Christmas Buddy Sitters"),
    ("坐姿小人", "Sitting Mini Figures"),
    ("天气小人", "Weather Mini Figures"),
    ("水果小人", "Fruit Mini Figures"),
    ("南瓜伙伴", "Pumpkin Buddies"),
    ("多色恐龙", "Multicolor Dinosaurs"),
    ("山海经", "Shanhai Myth Figures"),
    ("磁吸壁虎", "Magnetic Gecko"),
    ("猴子冰箱贴", "Monkey Fridge Magnet"),
]


def slugify(name: str, idx: int) -> str:
    base = re.sub(r"[^a-zA-Z0-9]+", "-", name.lower()).strip("-")
    if not base or len(base) < 2:
        base = f"toy-{idx:03d}"
    return f"jq-{idx:03d}-{base[:40]}"


def to_english(zh_name: str) -> str:
    for zh, en in WORD_MAP:
        if zh_name == zh or zh_name.startswith(zh):
            return en
    # fallback light mapping
    t = zh_name
    reps = [
        ("钥匙扣", " Keychain"),
        ("冰箱贴", " Fridge Magnet"),
        ("解压", " Fidget"),
        ("龙蛋", " Dragon Egg"),
        ("恐龙", " Dinosaur"),
        ("龙", " Dragon"),
        ("套装", " Set"),
        ("系列", " Series"),
        ("盒装", " Box Set"),
        ("夜光", " Glow"),
        ("多色", " Multicolor"),
        ("迷你", " Mini"),
        ("磁吸", " Magnetic"),
    ]
    for a, b in reps:
        t = t.replace(a, b)
    t = re.sub(r"\s+", " ", t).strip(" -")
    return t if re.search(r"[A-Za-z]", t) else f"3D Printed Toy — {zh_name}"


def parse_price_cny(raw) -> float | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or "线下" in s or "定做" in s or "单独" in s and "元" not in s:
        # try extract numbers anyway
        pass
    nums = re.findall(r"\d+(?:\.\d+)?", s.replace("，", ","))
    if not nums:
        return None
    vals = [float(n) for n in nums if float(n) < 500]
    if not vals:
        return None
    if len(vals) >= 2 and ("-" in s or "/" in s or "单色" in s or "多色" in s):
        return sum(vals[:2]) / 2
    return vals[0]


def retail_usd(cny: float) -> float:
    usd = cny * MARKUP / FX
    # Amazon-like .99 endings, floor at 4.99
    rounded = max(4.99, round(usd) - 0.01)
    if rounded < 5:
        return 4.99
    return round(rounded, 2)


def image_row(img) -> int | None:
    a = img.anchor
    if isinstance(a, str):
        from openpyxl.utils.cell import coordinate_from_string

        _, row = coordinate_from_string(a)
        return row
    try:
        return a._from.row + 1
    except Exception:
        return None


def save_image(img, dest: Path) -> bool:
    try:
        data = img._data()
        if callable(data):
            data = data()
        if not data:
            return False
        im = Image.open(BytesIO(data))
        im = im.convert("RGB")
        w, h = im.size
        max_side = 1100
        if max(w, h) > max_side:
            scale = max_side / max(w, h)
            im = im.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
        dest.parent.mkdir(parents=True, exist_ok=True)
        im.save(dest, "JPEG", quality=82, optimize=True)
        return True
    except Exception as e:
        print(f"  skip image {dest.name}: {e}")
        return False


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing xlsx: {XLSX}")

    if OUT_IMG.exists():
        shutil.rmtree(OUT_IMG)
    OUT_IMG.mkdir(parents=True, exist_ok=True)

    print("Loading workbook (large file)...")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Sheet1"]

    # Map row -> list of image indices
    images = list(ws._images)
    row_to_imgs: dict[int, list[int]] = {}
    for i, img in enumerate(images):
        r = image_row(img)
        if r:
            row_to_imgs.setdefault(r, []).append(i)

    # Build products
    products_raw: list[dict] = []
    current = None
    for row_i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=10, values_only=True), start=3):
        name, _photo, color, spec, pack, carton, qty, weight, price, note = row
        if name and str(name).strip():
            current = {
                "name_zh": str(name).strip(),
                "header_row": row_i,
                "rows": [],
                "image_rows": set(),
            }
            products_raw.append(current)
        if not current:
            continue
        current["rows"].append(
            {
                "row": row_i,
                "color": str(color).strip() if color else "",
                "spec": str(spec).strip() if spec and not str(spec).startswith("=_xlfn") else "",
                "pack": str(pack).strip() if pack else "",
                "carton": str(carton).strip() if carton else "",
                "qty": str(qty).strip() if qty else "",
                "weight": str(weight).strip() if weight else "",
                "price_raw": str(price).strip() if price else "",
                "price_cny": parse_price_cny(price),
                "note": str(note).strip() if note else "",
            }
        )
        if row_i in row_to_imgs:
            current["image_rows"].add(row_i)

    catalog = []
    used_img_idxs: set[int] = set()

    for idx, raw in enumerate(products_raw, start=1):
        name_zh = raw["name_zh"]
        name_en = to_english(name_zh)
        pid = slugify(name_en, idx)

        # pick images: prefer header row, then any rows for this product
        candidate_rows = [raw["header_row"]] + sorted(raw["image_rows"])
        img_paths = []
        for r in candidate_rows:
            for img_i in row_to_imgs.get(r, []):
                if img_i in used_img_idxs:
                    continue
                dest = OUT_IMG / f"{pid}-{len(img_paths)+1}.jpg"
                if save_image(images[img_i], dest):
                    used_img_idxs.add(img_i)
                    img_paths.append(f"/assets/toys/jinqi/{dest.name}")
                if len(img_paths) >= 4:
                    break
            if len(img_paths) >= 4:
                break

        # If still no image, grab next unused image near header row
        if not img_paths:
            nearby = sorted(
                ((abs(r - raw["header_row"]), r) for r in row_to_imgs),
                key=lambda x: x[0],
            )
            for _, r in nearby[:8]:
                for img_i in row_to_imgs[r]:
                    if img_i in used_img_idxs:
                        continue
                    dest = OUT_IMG / f"{pid}-1.jpg"
                    if save_image(images[img_i], dest):
                        used_img_idxs.add(img_i)
                        img_paths.append(f"/assets/toys/jinqi/{dest.name}")
                        break
                if img_paths:
                    break

        variants = []
        prices_cny = []
        for v in raw["rows"]:
            if v["price_cny"] is not None:
                prices_cny.append(v["price_cny"])
            variants.append(
                {
                    "spec": v["spec"],
                    "color": v["color"],
                    "priceCny": v["price_cny"],
                    "priceUsd": retail_usd(v["price_cny"]) if v["price_cny"] is not None else None,
                    "weight": v["weight"],
                    "note": v["note"],
                }
            )

        if prices_cny:
            base_cny = min(prices_cny)
            price = retail_usd(base_cny)
            price_max = retail_usd(max(prices_cny))
        else:
            base_cny = None
            price = 9.99
            price_max = 9.99

        specs = [v["spec"] for v in variants if v["spec"]]
        notes = [v["note"] for v in variants if v["note"]]
        spec_text = " / ".join(specs[:6])
        note_text = notes[0] if notes else ""

        zh_desc_parts = []
        if spec_text:
            zh_desc_parts.append(f"规格：{spec_text}")
        if note_text:
            zh_desc_parts.append(note_text)
        if len(prices_cny) > 1:
            zh_desc_parts.append(f"多尺寸可选，零售约 ${price:.2f}–${price_max:.2f}")
        else:
            zh_desc_parts.append("PLA 彩色3D打印玩具，可定制尺寸与颜色。")
        zh_desc = "。".join(zh_desc_parts)

        en_desc_parts = []
        if spec_text:
            en_desc_parts.append(f"Sizes: {spec_text}")
        if note_text:
            en_desc_parts.append(note_text)
        if len(prices_cny) > 1:
            en_desc_parts.append(f"Multiple sizes from ${price:.2f}")
        else:
            en_desc_parts.append("Color PLA 3D-printed toy. Size and color customizable.")
        en_desc = ". ".join(en_desc_parts)

        fallback_img = "/assets/toys/dragon.png"
        catalog.append(
            {
                "id": pid,
                "collection": "toys",
                "image": img_paths[0] if img_paths else fallback_img,
                "images": img_paths,
                "price": price,
                "priceMax": price_max,
                "costCny": base_cny,
                "material": "PLA",
                "supplier": "Jinqi",
                "variants": variants,
                "zh": {"name": name_zh, "desc": zh_desc},
                "en": {"name": name_en, "desc": en_desc},
            }
        )
        print(f"[{idx:03d}/{len(products_raw)}] {name_zh} → {pid} imgs={len(img_paths)} ${price}")

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")

    # Emit JS module (strip heavy variants detail to keep file reasonable — keep summary)
    js_items = []
    for p in catalog:
        slim = {
            "id": p["id"],
            "collection": p["collection"],
            "image": p["image"],
            "images": p["images"],
            "price": p["price"],
            "priceMax": p["priceMax"],
            "material": p["material"],
            "zh": p["zh"],
            "en": p["en"],
        }
        js_items.append(slim)

    OUT_JS.write_text(
        "// Auto-generated from 今奇.xlsx — do not edit by hand.\n"
        f"export const toyProducts = {json.dumps(js_items, ensure_ascii=False, indent=2)};\n",
        encoding="utf-8",
    )

    total_bytes = sum(f.stat().st_size for f in OUT_IMG.glob("*.jpg"))
    print(
        f"\nDone: {len(catalog)} products, {len(list(OUT_IMG.glob('*.jpg')))} images, "
        f"{total_bytes/1e6:.1f} MB → {OUT_JS}"
    )


if __name__ == "__main__":
    main()
